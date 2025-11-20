"""
Report Generation and Export Utilities
"""
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
import csv
import json


class ReportGenerator:
    """Base class for generating reports in various formats"""
    
    def __init__(self, title, data, user=None):
        self.title = title
        self.data = data
        self.user = user
        self.timestamp = timezone.now()
    
    def generate_pdf(self):
        """Generate PDF report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(self.title, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Metadata
        meta_style = styles['Normal']
        elements.append(Paragraph(f"Generated: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        if self.user:
            elements.append(Paragraph(f"User: {self.user.email}", meta_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Data table
        if isinstance(self.data, list) and len(self.data) > 0:
            # Create table from data
            table_data = [list(self.data[0].keys())]  # Headers
            for row in self.data:
                table_data.append(list(row.values()))
            
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def generate_excel(self):
        """Generate Excel report with charts"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = self.title
        title_cell.font = Font(size=18, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Metadata
        ws['A2'] = f"Generated: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
        if self.user:
            ws['A3'] = f"User: {self.user.email}"
        
        # Headers style
        header_fill = PatternFill(start_color='D0D7F7', end_color='D0D7F7', fill_type='solid')
        header_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Data
        if isinstance(self.data, list) and len(self.data) > 0:
            start_row = 5
            headers = list(self.data[0].keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row_data in enumerate(self.data, start_row + 1):
                for col_idx, value in enumerate(row_data.values(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Alternate row colors
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generate_csv(self):
        """Generate CSV report"""
        output = BytesIO()
        writer = csv.writer(output)
        
        # Title and metadata
        writer.writerow([self.title])
        writer.writerow([f"Generated: {self.timestamp.strftime('%Y-%m-%d %H:%M:%S')}"])
        if self.user:
            writer.writerow([f"User: {self.user.email}"])
        writer.writerow([])  # Empty row
        
        # Data
        if isinstance(self.data, list) and len(self.data) > 0:
            # Headers
            writer.writerow(self.data[0].keys())
            # Rows
            for row in self.data:
                writer.writerow(row.values())
        
        output.seek(0)
        return output
    
    def generate_json(self):
        """Generate JSON report"""
        report = {
            'title': self.title,
            'generated_at': self.timestamp.isoformat(),
            'user': self.user.email if self.user else None,
            'data': self.data
        }
        return json.dumps(report, indent=2, default=str)


class TaskReportGenerator(ReportGenerator):
    """Specialized report generator for tasks"""
    
    def prepare_task_data(self, tasks):
        """Prepare task data for reporting"""
        data = []
        for task in tasks:
            data.append({
                'ID': task.id,
                'Title': task.title,
                'Status': task.get_status_display(),
                'Priority': task.get_priority_display() if hasattr(task, 'priority') else 'N/A',
                'Project': task.project.name if task.project else 'N/A',
                'Assigned To': task.assigned_to.email if task.assigned_to else 'Unassigned',
                'Due Date': task.due_date.strftime('%Y-%m-%d') if task.due_date else 'N/A',
                'Created': task.created_at.strftime('%Y-%m-%d'),
            })
        return data


class TimeTrackingReportGenerator(ReportGenerator):
    """Specialized report generator for time tracking"""
    
    def prepare_time_data(self, time_entries):
        """Prepare time tracking data for reporting"""
        data = []
        total_minutes = 0
        
        for entry in time_entries:
            duration = entry.duration_minutes or 0
            total_minutes += duration
            
            data.append({
                'Date': entry.start_time.strftime('%Y-%m-%d'),
                'Task': entry.task.title,
                'User': entry.user.email,
                'Duration (hours)': f"{duration / 60:.2f}",
                'Billable': 'Yes' if entry.is_billable else 'No',
                'Description': entry.description or 'N/A',
            })
        
        # Add summary row
        data.append({
            'Date': 'TOTAL',
            'Task': '',
            'User': '',
            'Duration (hours)': f"{total_minutes / 60:.2f}",
            'Billable': '',
            'Description': '',
        })
        
        return data


class AnalyticsReportGenerator(ReportGenerator):
    """Specialized report generator for analytics"""
    
    def generate_dashboard_report(self, stats):
        """Generate comprehensive dashboard report"""
        data = [
            {'Metric': 'Total Tasks', 'Value': stats.get('total_tasks', 0)},
            {'Metric': 'Completed Tasks', 'Value': stats.get('completed_tasks', 0)},
            {'Metric': 'In Progress', 'Value': stats.get('in_progress', 0)},
            {'Metric': 'Overdue Tasks', 'Value': stats.get('overdue_tasks', 0)},
            {'Metric': 'Total Projects', 'Value': stats.get('total_projects', 0)},
            {'Metric': 'Active Projects', 'Value': stats.get('active_projects', 0)},
            {'Metric': 'Team Members', 'Value': stats.get('team_members', 0)},
            {'Metric': 'Completion Rate', 'Value': f"{stats.get('completion_rate', 0)}%"},
        ]
        return data


def export_report(report_type, format, data, user, filename_prefix):
    """
    Export report in specified format.
    
    Args:
        report_type: Type of report (task, time, analytics)
        format: Export format (pdf, excel, csv, json)
        data: Data to export
        user: User requesting the report
        filename_prefix: Prefix for the filename
    
    Returns:
        HttpResponse with the generated file
    """
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{filename_prefix}_{timestamp}"
    
    # Select appropriate generator
    if report_type == 'task':
        generator = TaskReportGenerator(f"{filename_prefix} Report", [], user)
        processed_data = generator.prepare_task_data(data)
    elif report_type == 'time':
        generator = TimeTrackingReportGenerator(f"{filename_prefix} Report", [], user)
        processed_data = generator.prepare_time_data(data)
    elif report_type == 'analytics':
        generator = AnalyticsReportGenerator(f"{filename_prefix} Report", data, user)
        processed_data = data
    else:
        generator = ReportGenerator(f"{filename_prefix} Report", data, user)
        processed_data = data
    
    generator.data = processed_data
    
    # Generate in requested format
    if format == 'pdf':
        buffer = generator.generate_pdf()
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    elif format == 'excel':
        buffer = generator.generate_excel()
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    elif format == 'csv':
        buffer = generator.generate_csv()
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    elif format == 'json':
        json_data = generator.generate_json()
        response = HttpResponse(json_data, content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
    else:
        raise ValueError(f"Unsupported format: {format}")
    
    return response

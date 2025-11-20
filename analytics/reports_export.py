"""
Report generation for analytics.
Supports PDF and Excel export formats.
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from tasks.models import Task
from projects.models import Project
from users.models import User
import io


def generate_pdf_report(user, date_from=None, date_to=None):
    """
    Generate PDF report for user's tasks and projects.
    
    Args:
        user: User instance
        date_from: Start date (optional)
        date_to: End date (optional)
    
    Returns:
        HttpResponse with PDF content
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    title = Paragraph("Task Management Report", title_style)
    elements.append(title)
    
    # Date range
    if date_from and date_to:
        date_range = f"Period: {date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    else:
        date_range = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    date_para = Paragraph(date_range, styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 20))
    
    # User info
    user_info = f"Report for: {user.get_full_name() or user.email}"
    elements.append(Paragraph(user_info, styles['Normal']))
    elements.append(Spacer(1, 30))
    
    # Get data
    tasks_query = Task.objects.filter(assigned_to=user)
    if date_from:
        tasks_query = tasks_query.filter(created_at__gte=date_from)
    if date_to:
        tasks_query = tasks_query.filter(created_at__lte=date_to)
    
    tasks = tasks_query.select_related('project')
    
    # Summary Statistics
    elements.append(Paragraph("Summary Statistics", heading_style))
    
    stats_data = [
        ['Metric', 'Count'],
        ['Total Tasks', tasks.count()],
        ['Completed Tasks', tasks.filter(status='done').count()],
        ['In Progress Tasks', tasks.filter(status='in_progress').count()],
        ['Pending Tasks', tasks.filter(status='todo').count()],
        ['High Priority', tasks.filter(priority='high').count()],
        ['Overdue Tasks', tasks.filter(due_date__lt=timezone.now().date(), status__in=['todo', 'in_progress']).count()],
    ]
    
    stats_table = Table(stats_data, colWidths=[4*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06B6D4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 30))
    
    # Task List
    elements.append(Paragraph("Task Details", heading_style))
    
    if tasks.exists():
        task_data = [['Title', 'Project', 'Status', 'Priority', 'Due Date']]
        
        for task in tasks[:50]:  # Limit to 50 tasks
            task_data.append([
                task.title[:40] + '...' if len(task.title) > 40 else task.title,
                task.project.name[:30] + '...' if len(task.project.name) > 30 else task.project.name,
                task.get_status_display(),
                task.get_priority_display(),
                task.due_date.strftime('%m/%d/%Y') if task.due_date else 'No due date'
            ])
        
        task_table = Table(task_data, colWidths=[2.5*inch, 1.8*inch, 1*inch, 0.8*inch, 1*inch])
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        
        elements.append(task_table)
    else:
        elements.append(Paragraph("No tasks found for this period.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF content
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response


def generate_excel_report(user, date_from=None, date_to=None):
    """
    Generate Excel report for user's tasks and projects.
    
    Args:
        user: User instance
        date_from: Start date (optional)
        date_to: End date (optional)
    
    Returns:
        HttpResponse with Excel content
    """
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = "Task Management Report"
    ws_summary['A1'].font = Font(size=18, bold=True, color="1F2937")
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:E1')
    
    # Date info
    if date_from and date_to:
        date_range = f"Period: {date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    else:
        date_range = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    ws_summary['A2'] = date_range
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A2:E2')
    
    # User info
    ws_summary['A3'] = f"Report for: {user.get_full_name() or user.email}"
    ws_summary['A3'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A3:E3')
    
    # Get data
    tasks_query = Task.objects.filter(assigned_to=user)
    if date_from:
        tasks_query = tasks_query.filter(created_at__gte=date_from)
    if date_to:
        tasks_query = tasks_query.filter(created_at__lte=date_to)
    
    tasks = tasks_query.select_related('project')
    
    # Statistics
    ws_summary['A5'] = "Metric"
    ws_summary['B5'] = "Count"
    ws_summary['A5'].font = Font(bold=True)
    ws_summary['B5'].font = Font(bold=True)
    ws_summary['A5'].fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    ws_summary['B5'].fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    
    stats = [
        ('Total Tasks', tasks.count()),
        ('Completed Tasks', tasks.filter(status='done').count()),
        ('In Progress Tasks', tasks.filter(status='in_progress').count()),
        ('Pending Tasks', tasks.filter(status='todo').count()),
        ('High Priority', tasks.filter(priority='high').count()),
        ('Overdue Tasks', tasks.filter(due_date__lt=timezone.now().date(), status__in=['todo', 'in_progress']).count()),
    ]
    
    row = 6
    for stat_name, stat_value in stats:
        ws_summary[f'A{row}'] = stat_name
        ws_summary[f'B{row}'] = stat_value
        row += 1
    
    # Apply borders
    for row in range(5, row):
        for col in ['A', 'B']:
            cell = ws_summary[f'{col}{row}']
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Task Details Sheet
    ws_tasks = wb.create_sheet("Tasks")
    
    # Headers
    headers = ['Title', 'Project', 'Status', 'Priority', 'Due Date', 'Created Date', 'Description']
    for col_num, header in enumerate(headers, 1):
        cell = ws_tasks.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, task in enumerate(tasks, 2):
        ws_tasks.cell(row=row_num, column=1, value=task.title)
        ws_tasks.cell(row=row_num, column=2, value=task.project.name)
        ws_tasks.cell(row=row_num, column=3, value=task.get_status_display())
        ws_tasks.cell(row=row_num, column=4, value=task.get_priority_display())
        ws_tasks.cell(row=row_num, column=5, value=task.due_date.strftime('%m/%d/%Y') if task.due_date else 'No due date')
        ws_tasks.cell(row=row_num, column=6, value=task.created_at.strftime('%m/%d/%Y'))
        ws_tasks.cell(row=row_num, column=7, value=task.description[:100] if task.description else '')
    
    # Auto-size columns
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws_tasks[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_tasks.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response
